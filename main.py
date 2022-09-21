import datetime
import json
import os
import re

from openpyxl import load_workbook


# created by huzongyao


class IOSAppDeploy:
    SITE_ROOT = 'https://cop-dev.github.io/ios-apps/plist/'
    DIST_PATH = './dist/plist/'
    EXCEL_PATH = './apps.xlsx'
    PLIST_TPL_PATH = './tpl/plist.xml'

    def __init__(self):
        self.plist_tpl = self._load_doc_tpl(self.PLIST_TPL_PATH)

    @staticmethod
    def check_file_name(name=None):
        if name is None:
            print("name is None!")
            return
        reg = re.compile(r'[\\/:*?"<>|\r\n\t]+')
        valid_name = reg.findall(name)
        if valid_name:
            for nv in valid_name:
                name = name.replace(nv, "_")
        return name

    @staticmethod
    def _load_doc_tpl(doc_path):
        with open(doc_path, 'r', encoding='utf-8') as doc:
            txt = doc.read()
        return txt

    def generate_some_plist(self, idx, row, out_path):
        name = row[0].value
        bundle_id = row[1].value
        version = row[2].value
        bundle_url = row[3].value
        plist_name = '%s_%s_%s.plist' % (bundle_id, version, idx)
        plist_path = '%s%s' % (out_path, plist_name)
        txt = self.plist_tpl
        txt = txt.replace('%app_name%', name)
        txt = txt.replace('%bundle_version%', version)
        txt = txt.replace('%bundle_id%', bundle_id)
        txt = txt.replace('%bundle_url%', bundle_url)
        with open(plist_path, 'w', encoding='utf-8') as f:
            f.write(txt)
        return plist_name

    def generate_single_sheet(self, sheet_title, sheet, out_path):
        if not os.path.exists(out_path):
            os.makedirs(out_path)
        js_obj = {'name': sheet_title}
        js_list = []
        for idx, row in enumerate(sheet.rows):
            if idx == 0:
                continue
            name = row[0].value
            bundle_id = row[1].value
            version = row[2].value
            plist_name = self.generate_some_plist(idx, row, out_path)
            plist_path = '%s%s/%s' % (self.SITE_ROOT, sheet_title, plist_name)
            href = 'itms-services://?action=download-manifest&url=%s' % plist_path
            js_list.append(
                {'category': sheet_title, 'name': name, 'version': version, 'href': href, 'bundle_id': bundle_id})
        js_obj['apps'] = js_list
        js_path = '%s/index.json' % out_path
        with open(js_path, 'w', encoding='utf-8') as f:
            json.dump(js_obj, f)

    def page_generate_work(self):
        if not os.path.exists(self.EXCEL_PATH):
            print('EXCEL File Not Exists!!')
            return
        if not os.path.exists(self.DIST_PATH):
            os.makedirs(self.DIST_PATH)
        wb = load_workbook(self.EXCEL_PATH)
        time_txt = datetime.datetime.now().strftime("%Y%m%d %H%M%S")
        js_obj = {'time': time_txt}
        js_list = []
        for sheet in wb:
            sheet_title = self.check_file_name(sheet.title)
            sheet_path = '%s%s/' % (self.DIST_PATH, sheet_title)
            self.generate_single_sheet(sheet_title, sheet, sheet_path)
            js_list.append({'name': sheet_title, 'count': sheet.max_row - 1})
        js_obj['list'] = js_list
        js_path = '%sindex.json' % self.DIST_PATH
        with open(js_path, 'w', encoding='utf-8') as f:
            json.dump(js_obj, f)


if __name__ == "__main__":
    dep = IOSAppDeploy()
    dep.page_generate_work()
